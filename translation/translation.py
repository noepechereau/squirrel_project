import openpyxl
import json


def load_translations():
    workbook = openpyxl.load_workbook("workbooks/languages.xlsx")
    worksheet = workbook.active
    json_object = {}
    for i in range(3, 1000):
        if worksheet.cell(row=i, column=2).value is None or worksheet.cell(row=i, column=3).value is None:
            continue
        json_object[worksheet.cell(row=i, column=3).value] = worksheet.cell(row=i, column=2).value
    workbook.close()
    file = open("translation/translations.json", "w")
    file.write(json.dumps(json_object, ensure_ascii=False))
    file.close()


def translate(name):
    with open("translation/translations.json", "r") as file:
        content = file.read()
    try:
        res = json.loads(content)[name]
    except:
        return ""
    return res
